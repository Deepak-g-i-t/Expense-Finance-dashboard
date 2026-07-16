import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from pathlib import Path
from datetime import datetime
import models
import os

HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
ALT_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
BORDER = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)


def style_header(ws, headers):
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_num)].width = max(15, len(header) + 5)
    ws.row_dimensions[1].height = 25


def style_row(ws, row_num, num_cols):
    fill = ALT_FILL if row_num % 2 == 0 else None
    for col_num in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col_num)
        if fill:
            cell.fill = fill
        cell.border = BORDER
        cell.alignment = Alignment(vertical='center')


def sync_excel_backup(db: Session, account: models.Account, backup_path: str):
    """Sync all data to Excel workbook backup."""
    Path(backup_path).parent.mkdir(parents=True, exist_ok=True)
    
    wb = openpyxl.Workbook()
    
    # ── Expenses Sheet ────────────────────────────────────────────────────────
    ws_exp = wb.active
    ws_exp.title = "Expenses"
    exp_headers = ["ID", "Date", "Time", "Amount (₹)", "Category", "Description", "Merchant", "Payment Method", "Notes", "Created At"]
    style_header(ws_exp, exp_headers)
    
    expenses = db.query(models.Expense).filter(models.Expense.account_id == account.id).order_by(models.Expense.date.desc()).all()
    for i, exp in enumerate(expenses, 2):
        ws_exp.cell(row=i, column=1, value=exp.id)
        ws_exp.cell(row=i, column=2, value=exp.date)
        ws_exp.cell(row=i, column=3, value=exp.time)
        ws_exp.cell(row=i, column=4, value=exp.amount)
        ws_exp.cell(row=i, column=5, value=exp.category)
        ws_exp.cell(row=i, column=6, value=exp.description)
        ws_exp.cell(row=i, column=7, value=exp.merchant or "")
        ws_exp.cell(row=i, column=8, value=exp.payment_method)
        ws_exp.cell(row=i, column=9, value=exp.notes or "")
        ws_exp.cell(row=i, column=10, value=str(exp.created_at) if exp.created_at else "")
        style_row(ws_exp, i, len(exp_headers))

    # ── Income Sheet ──────────────────────────────────────────────────────────
    ws_inc = wb.create_sheet("Income")
    inc_headers = ["ID", "Date", "Time", "Amount (₹)", "Source", "Description", "Notes", "Created At"]
    style_header(ws_inc, inc_headers)
    
    incomes = db.query(models.Income).filter(models.Income.account_id == account.id).order_by(models.Income.date.desc()).all()
    for i, inc in enumerate(incomes, 2):
        ws_inc.cell(row=i, column=1, value=inc.id)
        ws_inc.cell(row=i, column=2, value=inc.date)
        ws_inc.cell(row=i, column=3, value=inc.time)
        ws_inc.cell(row=i, column=4, value=inc.amount)
        ws_inc.cell(row=i, column=5, value=inc.source)
        ws_inc.cell(row=i, column=6, value=inc.description)
        ws_inc.cell(row=i, column=7, value=inc.notes or "")
        ws_inc.cell(row=i, column=8, value=str(inc.created_at) if inc.created_at else "")
        style_row(ws_inc, i, len(inc_headers))

    # ── Budgets Sheet ─────────────────────────────────────────────────────────
    ws_bud = wb.create_sheet("Budgets")
    bud_headers = ["ID", "Category", "Budget (₹)", "Month", "Year", "Created At"]
    style_header(ws_bud, bud_headers)
    
    budgets = db.query(models.Budget).filter(models.Budget.account_id == account.id).all()
    for i, bud in enumerate(budgets, 2):
        ws_bud.cell(row=i, column=1, value=bud.id)
        ws_bud.cell(row=i, column=2, value=bud.category)
        ws_bud.cell(row=i, column=3, value=bud.amount)
        ws_bud.cell(row=i, column=4, value=bud.month)
        ws_bud.cell(row=i, column=5, value=bud.year)
        ws_bud.cell(row=i, column=6, value=str(bud.created_at) if bud.created_at else "")
        style_row(ws_bud, i, len(bud_headers))

    # ── Reminders Sheet ───────────────────────────────────────────────────────
    ws_rem = wb.create_sheet("Reminders")
    rem_headers = ["ID", "Title", "Description", "Amount (₹)", "Due Date", "Reminder Date", "Priority", "Repeat", "Status"]
    style_header(ws_rem, rem_headers)
    
    reminders = db.query(models.Reminder).filter(models.Reminder.account_id == account.id).all()
    for i, rem in enumerate(reminders, 2):
        ws_rem.cell(row=i, column=1, value=rem.id)
        ws_rem.cell(row=i, column=2, value=rem.title)
        ws_rem.cell(row=i, column=3, value=rem.description or "")
        ws_rem.cell(row=i, column=4, value=rem.amount or "")
        ws_rem.cell(row=i, column=5, value=rem.due_date)
        ws_rem.cell(row=i, column=6, value=rem.reminder_date)
        ws_rem.cell(row=i, column=7, value=rem.priority)
        ws_rem.cell(row=i, column=8, value=rem.repeat)
        ws_rem.cell(row=i, column=9, value=rem.status)
        style_row(ws_rem, i, len(rem_headers))

    # ── Categories Sheet ──────────────────────────────────────────────────────
    ws_cat = wb.create_sheet("Categories")
    cat_headers = ["ID", "Name", "Icon", "Color", "Is Default"]
    style_header(ws_cat, cat_headers)
    
    categories = db.query(models.Category).filter(models.Category.account_id == account.id).all()
    for i, cat in enumerate(categories, 2):
        ws_cat.cell(row=i, column=1, value=cat.id)
        ws_cat.cell(row=i, column=2, value=cat.name)
        ws_cat.cell(row=i, column=3, value=cat.icon)
        ws_cat.cell(row=i, column=4, value=cat.color)
        ws_cat.cell(row=i, column=5, value="Yes" if cat.is_default else "No")
        style_row(ws_cat, i, len(cat_headers))

    wb.save(backup_path)
    return backup_path
