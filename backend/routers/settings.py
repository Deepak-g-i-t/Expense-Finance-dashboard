from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import List
from database import get_db
from config import get_settings
import models, schemas, security
from excel_sync import sync_excel_backup
import openpyxl
import csv
import io
import os
from pathlib import Path
from datetime import datetime

router = APIRouter(prefix="/settings", tags=["Settings"])
settings_conf = get_settings()


@router.get("", response_model=schemas.AccountResponse)
def get_settings_data(
    current_account: models.Account = Depends(security.get_current_account)
):
    return current_account


@router.put("", response_model=schemas.AccountResponse)
def update_settings(
    payload: schemas.SettingsUpdate,
    current_account: models.Account = Depends(security.get_current_account),
    db: Session = Depends(get_db)
):
    for field, value in payload.dict(exclude_unset=True).items():
        setattr(current_account, field, value)
    db.commit()
    db.refresh(current_account)
    return current_account


@router.post("/change-pin")
def change_pin(
    payload: schemas.PinChange,
    current_account: models.Account = Depends(security.get_current_account),
    db: Session = Depends(get_db)
):
    if not security.verify_pin(payload.current_pin, current_account.hashed_pin):
        raise HTTPException(status_code=400, detail="Current PIN is incorrect")
    current_account.hashed_pin = security.hash_pin(payload.new_pin)
    db.commit()
    return {"message": "PIN changed successfully"}


@router.get("/export/excel")
def export_excel(
    current_account: models.Account = Depends(security.get_current_account),
    db: Session = Depends(get_db)
):
    export_path = f"./data/export_{current_account.account_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    sync_excel_backup(db, current_account, export_path)
    return FileResponse(
        export_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"expense_analytics_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )


@router.get("/export/csv")
def export_csv(
    current_account: models.Account = Depends(security.get_current_account),
    db: Session = Depends(get_db)
):
    expenses = db.query(models.Expense).filter(models.Expense.account_id == current_account.id).all()
    incomes = db.query(models.Income).filter(models.Income.account_id == current_account.id).all()

    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow(["Type", "Date", "Time", "Amount", "Category/Source", "Description", "Payment Method", "Merchant", "Notes"])
    for e in expenses:
        writer.writerow(["Expense", e.date, e.time, e.amount, e.category, e.description, e.payment_method, e.merchant or "", e.notes or ""])
    for i in incomes:
        writer.writerow(["Income", i.date, i.time, i.amount, i.source, i.description, "", "", i.notes or ""])

    csv_path = f"./data/export_{current_account.account_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    Path("./data").mkdir(exist_ok=True)
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        f.write(output.getvalue())

    return FileResponse(csv_path, media_type="text/csv", filename=f"expense_analytics_{datetime.now().strftime('%Y%m%d')}.csv")


@router.post("/import/excel")
def import_excel(
    file: UploadFile = File(...),
    current_account: models.Account = Depends(security.get_current_account),
    db: Session = Depends(get_db)
):
    content = file.file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    imported = {"expenses": 0, "incomes": 0, "skipped": 0}

    if "Expenses" in wb.sheetnames:
        ws = wb["Expenses"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            _, date, time, amount, category, description, merchant, payment_method, notes, _ = (list(row) + [None]*10)[:10]
            if not all([date, amount, category, description]):
                imported["skipped"] += 1
                continue
            # Dedup check
            exists = db.query(models.Expense).filter(
                models.Expense.account_id == current_account.id,
                models.Expense.date == str(date),
                models.Expense.amount == float(amount),
                models.Expense.description == str(description)
            ).first()
            if exists:
                imported["skipped"] += 1
                continue
            db.add(models.Expense(
                account_id=current_account.id,
                amount=float(amount), category=str(category),
                description=str(description), merchant=str(merchant) if merchant else None,
                date=str(date), time=str(time) if time else "00:00",
                payment_method=str(payment_method) if payment_method else "Cash",
                notes=str(notes) if notes else None
            ))
            imported["expenses"] += 1

    if "Income" in wb.sheetnames:
        ws = wb["Income"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            _, date, time, amount, source, description, notes, _ = (list(row) + [None]*8)[:8]
            if not all([date, amount, source, description]):
                imported["skipped"] += 1
                continue
            exists = db.query(models.Income).filter(
                models.Income.account_id == current_account.id,
                models.Income.date == str(date),
                models.Income.amount == float(amount),
                models.Income.description == str(description)
            ).first()
            if exists:
                imported["skipped"] += 1
                continue
            db.add(models.Income(
                account_id=current_account.id,
                amount=float(amount), source=str(source),
                description=str(description), date=str(date),
                time=str(time) if time else "00:00",
                notes=str(notes) if notes else None
            ))
            imported["incomes"] += 1

    db.commit()
    sync_excel_backup(db, current_account, settings_conf.EXCEL_BACKUP_PATH)
    return imported
